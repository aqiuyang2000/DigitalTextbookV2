# FILE: _mixin_export_single_page_hotspots_to_excel.py
#
# 功能: 为 ExcelProcessor 添加导出单个页面热区到 Excel 的功能

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


class ExcelProcessorExportSinglePageHotspotsToExcelMixin:
    """
    为 ExcelProcessor 添加导出单个页面热区到 Excel 文件的功能。
    """

    @staticmethod
    def export_single_page_hotspots_to_excel(page_data: dict, file_path: str):
        """
        将单个页面的热区数据导出到 Excel 文件。

        Args:
            page_data (dict): 页面数据，包含 hotspots 列表
            file_path (str): 导出的 Excel 文件路径
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"Page {page_data.get('page_index', 1)}"

        # 设置表头
        headers = ["id", "绘制形状", "x", "y", "宽", "高", "类型", "链接/文件", "打开方式", "说明"]
        sheet.append(headers)

        # 写入热区数据
        for hotspot in page_data.get('hotspots', []):
            data = hotspot.get('data', {})
            pos = hotspot.get('pos', {})
            rect = hotspot.get('rect', {})

            hotspot_type = data.get('hotspot_type', 'url')
            link_or_file = ""
            target = ""

            if hotspot_type == 'url':
                link_or_file = data.get('url_data', {}).get('url', '')
                target = data.get('url_data', {}).get('target', '')
            elif hotspot_type == 'file':
                link_or_file = data.get('file_data', {}).get('source_path', '')
                target = data.get('file_data', {}).get('display', '')

            row_data = [
                data.get('id', 'N/A'),
                hotspot.get('type', 'rectangle'),
                pos.get('x', 0),
                pos.get('y', 0),
                rect.get('w', 0),
                rect.get('h', 0),
                hotspot_type,
                link_or_file,
                target,
                data.get('description', '')
            ]
            sheet.append(row_data)

        # 设置表头样式
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center')
        for cell in sheet[1]:
            cell.font = header_font
            cell.alignment = center_alignment

        # 调整列宽
        for i, column_cells in enumerate(sheet.columns, 1):
            max_length = 0
            column = get_column_letter(i)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[column].width = (max_length + 2) if max_length > 0 else 12

        # 保存文件
        workbook.save(file_path)