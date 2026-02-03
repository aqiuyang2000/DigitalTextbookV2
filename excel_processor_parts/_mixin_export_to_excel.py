# FILE: _mixin_export_to_excel.py
#
# 功能: 为 ExcelProcessor 添加导出大纲数据到 Excel 的功能

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


class ExcelProcessorExportToExcelMixin:
    """
    为 ExcelProcessor 添加导出大纲数据到 Excel 文件的功能。
    """

    @staticmethod
    def export_to_excel(outline_data: list, file_path: str):
        """
        将大纲数据导出到 Excel 文件。

        Args:
            outline_data (list): 大纲数据列表
            file_path (str): 导出的 Excel 文件路径
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Outline"

        # 设置表头
        headers = ["Level", "Title", "Page"]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        def _recursive_write(nodes, level):
            for node in nodes:
                try:
                    sheet.append([level, node['title'], node['page']])
                    if node.get('children'):
                        _recursive_write(node['children'], level + 1)
                except Exception as e:
                    print(f"警告：写入行时出错，已跳过 - Node: {node}, Error: {e}")

        # 递归写入数据
        _recursive_write(outline_data, 1)

        # 调整列宽
        for i, column_cells in enumerate(sheet.columns, 1):
            max_length = 0
            column = get_column_letter(i)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[column].width = (max_length + 2)

        # 保存文件
        workbook.save(file_path)