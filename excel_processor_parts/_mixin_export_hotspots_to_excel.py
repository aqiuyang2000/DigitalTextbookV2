# FILE: _mixin_export_hotspots_to_excel.py
#
# 功能: 为 ExcelProcessor 添加导出所有页面热区到 Excel 的功能

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


class ExcelProcessorExportHotspotsToExcelMixin:
    """
    为 ExcelProcessor 添加导出所有页面热区到 Excel 文件的功能。
    每个页面一个工作表。
    """

    @staticmethod
    def export_hotspots_to_excel(project_data: dict, file_path: str):
        """
        将整个项目的所有页面热区数据导出到 Excel 文件。
        每个页面一个工作表。

        Args:
            project_data (dict): 项目数据，包含 pages 列表
            file_path (str): 导出的 Excel 文件路径
        """
        workbook = openpyxl.Workbook()

        # 移除默认创建的 Sheet
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center')

        # 遍历每个页面
        for page_index, page in enumerate(project_data.get('pages', [])):
            sheet_name = f"Page {page_index + 1}"
            sheet = workbook.create_sheet(title=sheet_name)

            # 设置表头
            headers = ["id", "绘制形状", "x", "y", "宽", "高", "类型", "链接/文件", "打开方式", "说明"]
            sheet.append(headers)

            # 写入热区数据
            for hotspot in page.get('hotspots', []):
                data = hotspot.get('data', {})
                pos = hotspot.get('pos', {})
                rect = hotspot.get('rect', {})

                hotspot_type = data.get('hotspot_type', 'url')
                link_or_file = ""
                target = ""

                if hotspot_type == 'url':
                    link_or_file = data.get('url_data', {}).get('url', '')
                    target = data.get('url_data', {}).get('target', '')
                elif hotspot_type == 'file':
                    link_or_file = data.get('file_data', {}).get('source_path', '')
                    target = data.get('file_data', {}).get('display', '')

                row_data = [
                    data.get('id', 'N/A'),
                    hotspot.get('type', 'rectangle'),
                    pos.get('x', 0),
                    pos.get('y', 0),
                    rect.get('w', 0),
                    rect.get('h', 0),
                    hotspot_type,
                    link_or_file,
                    target,
                    data.get('description', '')
                ]
                sheet.append(row_data)

            # 设置表头样式
            for cell in sheet[1]:
                cell.font = header_font
                cell.alignment = center_alignment

            # 调整列宽
            for i, column_cells in enumerate(sheet.columns, 1):
                max_length = 0
                column = get_column_letter(i)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                sheet.column_dimensions[column].width = (max_length + 2) if max_length > 0 else 12

        # 如果没有数据，创建一个空的工作表
        if not workbook.sheetnames:
            workbook.create_sheet(title="No Data")

        # 保存文件
        workbook.save(file_path)